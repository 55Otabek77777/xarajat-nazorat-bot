from flask import Flask, jsonify, request, send_file, render_template_string, session, redirect
from flask_cors import CORS
import sqlite3
from datetime import datetime, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import requests
import secrets

app = Flask(__name__)
app.secret_key = secrets.token_hex(32)
CORS(app)

DB_PATH = "expenses.db"
ADMIN_PASSWORD = "@Mirzo77"

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def get_usd_rate():
    try:
        response = requests.get("https://cbu.uz/uz/arkhiv-kursov-valyut/json/", timeout=5)
        data = response.json()
        for currency in data:
            if currency["Ccy"] == "USD":
                return {
                    "rate": float(currency["Rate"]),
                    "date": currency["Date"],
                    "diff": currency.get("Diff", "0")
                }
    except:
        pass
    return {"rate": 12650, "date": datetime.now().strftime("%d.%m.%Y"), "diff": "0"}

@app.route("/")
def index():
    if not session.get("authenticated"):
        return render_template_string(LOGIN_TEMPLATE)
    return render_template_string(HTML_TEMPLATE)

@app.route("/login", methods=["POST"])
def login():
    password = request.form.get("password")
    if password == ADMIN_PASSWORD:
        session["authenticated"] = True
        session.permanent = True
        app.permanent_session_lifetime = timedelta(hours=24)
        return redirect("/")
    return render_template_string(LOGIN_TEMPLATE, error="Parol noto'g'ri!")

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")

@app.route("/api/currency")
def get_currency():
    if not session.get("authenticated"):
        return jsonify({"error": "Unauthorized"}), 401
    return jsonify(get_usd_rate())

@app.route("/api/expenses")
def get_expenses():
    if not session.get("authenticated"):
        return jsonify({"error": "Unauthorized"}), 401
    
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    
    conn = get_db()
    cursor = conn.cursor()
    
    if start_date and end_date:
        cursor.execute("""
            SELECT * FROM transactions 
            WHERE date BETWEEN ? AND ? 
            ORDER BY date DESC
        """, (start_date, end_date))
    else:
        cursor.execute("SELECT * FROM transactions ORDER BY date DESC LIMIT 100")
    
    rows = cursor.fetchall()
    conn.close()
    
    expenses = []
    for row in rows:
        expenses.append({
            "id": row["id"],
            "user_id": row["user_id"],
            "type": row["type"],
            "category": row["category"],
            "amount": row["amount"],
            "description": row["description"],
            "date": row["date"]
        })
    
    return jsonify(expenses)

@app.route("/api/stats")
def get_stats():
    if not session.get("authenticated"):
        return jsonify({"error": "Unauthorized"}), 401
    
    period = request.args.get("period", "all")
    
    conn = get_db()
    cursor = conn.cursor()
    
    if period == "today":
        date_filter = datetime.now().strftime("%Y-%m-%d")
        cursor.execute("SELECT SUM(amount) as total, COUNT(*) as count FROM transactions WHERE date LIKE ?", (f"{date_filter}%",))
    elif period == "week":
        week_ago = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
        cursor.execute("SELECT SUM(amount) as total, COUNT(*) as count FROM transactions WHERE date >= ?", (week_ago,))
    elif period == "month":
        month_start = datetime.now().replace(day=1).strftime("%Y-%m-%d")
        cursor.execute("SELECT SUM(amount) as total, COUNT(*) as count FROM transactions WHERE date >= ?", (month_start,))
    elif period == "year":
        year_start = datetime.now().replace(month=1, day=1).strftime("%Y-%m-%d")
        cursor.execute("SELECT SUM(amount) as total, COUNT(*) as count FROM transactions WHERE date >= ?", (year_start,))
    else:
        cursor.execute("SELECT SUM(amount) as total, COUNT(*) as count FROM transactions")
    
    result = cursor.fetchone()
    
    cursor.execute("SELECT category, SUM(amount) as total FROM transactions GROUP BY category ORDER BY total DESC LIMIT 10")
    categories = cursor.fetchall()
    
    conn.close()
    
    return jsonify({
        "total": result["total"] or 0,
        "count": result["count"] or 0,
        "categories": [{"name": c["category"], "amount": c["total"]} for c in categories]
    })

@app.route("/api/export")
def export_excel():
    if not session.get("authenticated"):
        return "Unauthorized", 401
    
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    
    conn = get_db()
    cursor = conn.cursor()
    
    if start_date and end_date:
        cursor.execute("""
            SELECT * FROM transactions 
            WHERE date BETWEEN ? AND ? 
            ORDER BY date DESC
        """, (start_date, end_date))
    else:
        cursor.execute("SELECT * FROM transactions ORDER BY date DESC")
    
    rows = cursor.fetchall()
    conn.close()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Xarajatlar"
    
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    headers = ["ID", "Sana", "Kategoriya", "Summa (som)", "Izoh"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    total = 0
    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row["id"])
        ws.cell(row=row_idx, column=2, value=row["date"])
        ws.cell(row=row_idx, column=3, value=row["category"])
        ws.cell(row=row_idx, column=4, value=row["amount"])
        ws.cell(row=row_idx, column=5, value=row["description"])
        total += row["amount"]
    
    ws.cell(row=len(rows)+2, column=3, value="JAMI:").font = Font(bold=True)
    ws.cell(row=len(rows)+2, column=4, value=total).font = Font(bold=True, size=14)
    
    for col in range(1, 6):
        ws.column_dimensions[chr(64+col)].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"xarajatlar_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )

LOGIN_TEMPLATE = """
<!DOCTYPE html>
<html lang="uz">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Admin Login - Xarajatlar Nazorat</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 20px;
        }
        
        .login-container {
            background: white;
            padding: 40px;
            border-radius: 20px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            max-width: 400px;
            width: 100%;
            animation: slideUp 0.5s ease-out;
        }
        
        @keyframes slideUp {
            from { opacity: 0; transform: translateY(30px); }
            to { opacity: 1; transform: translateY(0); }
        }
        
        .login-header {
            text-align: center;
            margin-bottom: 30px;
        }
        
        .login-header h1 {
            font-size: 2em;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            margin-bottom: 10px;
        }
        
        .login-header p {
            color: #666;
        }
        
        .form-group {
            margin-bottom: 20px;
        }
        
        label {
            display: block;
            margin-bottom: 8px;
            font-weight: 600;
            color: #333;
        }
        
        input[type="password"] {
            width: 100%;
            padding: 15px;
            border: 2px solid #ddd;
            border-radius: 8px;
            font-size: 1em;
            transition: all 0.3s;
        }
        
        input[type="password"]:focus {
            border-color: #667eea;
            outline: none;
        }
        
        button {
            width: 100%;
            padding: 15px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            border-radius: 8px;
            font-size: 1.1em;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.3s;
        }
        
        button:hover {
            transform: translateY(-2px);
            box-shadow: 0 5px 15px rgba(102, 126, 234, 0.4);
        }
        
        .error {
            background: #fee;
            color: #c33;
            padding: 10px;
            border-radius: 8px;
            margin-bottom: 20px;
            text-align: center;
        }
    </style>
</head>
<body>
    <div class="login-container">
        <div class="login-header">
            <h1>🔐 Admin Panel</h1>
            <p>Xarajatlar Nazorat Tizimi</p>
        </div>
        
        {% if error %}
        <div class="error">{{ error }}</div>
        {% endif %}
        
        <form method="POST" action="/login">
            <div class="form-group">
                <label for="password">Parol:</label>
                <input type="password" id="password" name="password" placeholder="Parolni kiriting" required autofocus>
            </div>
            
            <button type="submit">Kirish</button>
        </form>
    </div>
</body>
</html>
"""

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="uz">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Admin Panel - Xarajatlar Nazorat</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }
        
        .container {
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            border-radius: 20px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            overflow: hidden;
            animation: slideUp 0.5s ease-out;
        }
        
        @keyframes slideUp {
            from { opacity: 0; transform: translateY(30px); }
            to { opacity: 1; transform: translateY(0); }
        }
        
        .header {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            text-align: center;
            position: relative;
        }
        
        .header h1 {
            font-size: 2.5em;
            margin-bottom: 10px;
            text-shadow: 2px 2px 4px rgba(0,0,0,0.2);
        }
        
        .currency-widget {
            position: absolute;
            top: 30px;
            right: 30px;
            background: rgba(255,255,255,0.2);
            padding: 15px 25px;
            border-radius: 15px;
            backdrop-filter: blur(10px);
        }
        
        .currency-widget .rate {
            font-size: 1.8em;
            font-weight: bold;
        }
        
        .currency-widget .label {
            font-size: 0.9em;
            opacity: 0.9;
        }
        
        .logout-btn {
            position: absolute;
            top: 30px;
            left: 30px;
            background: rgba(255,255,255,0.2);
            padding: 10px 20px;
            border-radius: 8px;
            color: white;
            text-decoration: none;
            backdrop-filter: blur(10px);
        }
        
        .logout-btn:hover {
            background: rgba(255,255,255,0.3);
        }
        
        .filters {
            padding: 30px;
            background: #f8f9fa;
            border-bottom: 2px solid #e0e0e0;
            display: flex;
            gap: 15px;
            flex-wrap: wrap;
            align-items: center;
        }
        
        .filter-group {
            display: flex;
            gap: 10px;
            align-items: center;
        }
        
        label {
            font-weight: 600;
            color: #333;
        }
        
        input[type="date"], select, button {
            padding: 12px 20px;
            border: 2px solid #ddd;
            border-radius: 8px;
            font-size: 1em;
            transition: all 0.3s;
        }
        
        input[type="date"]:focus, select:focus {
            border-color: #667eea;
            outline: none;
        }
        
        button {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            cursor: pointer;
            font-weight: 600;
        }
        
        button:hover {
            transform: translateY(-2px);
            box-shadow: 0 5px 15px rgba(102, 126, 234, 0.4);
        }
        
        .stats {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
            gap: 20px;
            padding: 30px;
        }
        
        .stat-card {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 25px;
            border-radius: 15px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.1);
            animation: fadeIn 0.5s ease-out;
        }
        
        @keyframes fadeIn {
            from { opacity: 0; transform: scale(0.9); }
            to { opacity: 1; transform: scale(1); }
        }
        
        .stat-card h3 {
            font-size: 0.9em;
            opacity: 0.9;
            margin-bottom: 10px;
        }
        
        .stat-card .value {
            font-size: 2em;
            font-weight: bold;
        }
        
        .stat-card .usd {
            font-size: 1.2em;
            opacity: 0.8;
            margin-top: 5px;
        }
        
        .chart-container {
            padding: 30px;
            background: white;
        }
        
        .table-container {
            padding: 30px;
            overflow-x: auto;
        }
        
        table {
            width: 100%;
            border-collapse: collapse;
            background: white;
            box-shadow: 0 5px 15px rgba(0,0,0,0.1);
            border-radius: 10px;
            overflow: hidden;
        }
        
        thead {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
        }
        
        th, td {
            padding: 15px;
            text-align: left;
        }
        
        tbody tr:nth-child(even) {
            background: #f8f9fa;
        }
        
        tbody tr:hover {
            background: #e3f2fd;
            cursor: pointer;
            transform: scale(1.01);
            transition: all 0.2s;
        }
        
        .amount {
            font-weight: bold;
            color: #667eea;
        }
        
        .export-btn {
            margin: 30px;
            padding: 15px 40px;
            font-size: 1.1em;
            background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%);
        }
        
        .loading {
            text-align: center;
            padding: 50px;
            font-size: 1.5em;
            color: #667eea;
        }
        
        @media (max-width: 768px) {
            .header h1 { font-size: 1.8em; }
            .currency-widget, .logout-btn {
                position: static;
                margin-top: 15px;
            }
            .filters {
                flex-direction: column;
            }
            .filter-group {
                width: 100%;
            }
            input[type="date"], select, button {
                width: 100%;
            }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <a href="/logout" class="logout-btn">🚪 Chiqish</a>
            <h1>📊 Admin Panel</h1>
            <p>Xarajatlar Nazorat Tizimi</p>
            <div class="currency-widget">
                <div class="label">💵 1 USD</div>
                <div class="rate" id="usdRate">...</div>
            </div>
        </div>
        
        <div class="filters">
            <div class="filter-group">
                <label>Davr:</label>
                <select id="period" onchange="updatePeriod()">
                    <option value="today">Bugun</option>
                    <option value="week">Hafta</option>
                    <option value="month">Oy</option>
                    <option value="year">Yil</option>
                    <option value="custom">Tanlov</option>
                    <option value="all">Hammasi</option>
                </select>
            </div>
            
            <div class="filter-group" id="customDates" style="display:none;">
                <label>Dan:</label>
                <input type="date" id="startDate">
                <label>Gacha:</label>
                <input type="date" id="endDate">
            </div>
            
            <button onclick="loadData()">🔍 Qidirish</button>
            <button onclick="exportExcel()" class="export-btn">📥 Excel yuklash</button>
        </div>
        
        <div class="stats">
            <div class="stat-card">
                <h3>💰 JAMI SUMMA</h3>
                <div class="value" id="totalAmount">0 som</div>
                <div class="usd" id="totalUSD">$0.00</div>
            </div>
            <div class="stat-card">
                <h3>📦 XARAJATLAR SONI</h3>
                <div class="value" id="totalCount">0 ta</div>
            </div>
        </div>
        
        <div class="chart-container">
            <canvas id="categoryChart"></canvas>
        </div>
        
        <div class="table-container">
            <table id="expensesTable">
                <thead>
                    <tr>
                        <th>№</th>
                        <th>📅 Sana</th>
                        <th>📦 Kategoriya</th>
                        <th>💵 Summa (som)</th>
                        <th>💵 Summa (USD)</th>
                        <th>📝 Izoh</th>
                    </tr>
                </thead>
                <tbody id="tableBody">
                    <tr><td colspan="6" class="loading">Yuklanmoqda...</td></tr>
                </tbody>
            </table>
        </div>
    </div>
    
    <script>
        let chartInstance = null;
        let usdRate = 12650;
        
        async function loadCurrency() {
            try {
                const res = await fetch('/api/currency');
                const data = await res.json();
                usdRate = data.rate;
                document.getElementById('usdRate').textContent = formatNumber(data.rate) + ' som';
            } catch (e) {
                console.error('Currency error:', e);
            }
        }
        
        function updatePeriod() {
            const period = document.getElementById('period').value;
            const customDates = document.getElementById('customDates');
            
            if (period === 'custom') {
                customDates.style.display = 'flex';
            } else {
                customDates.style.display = 'none';
                loadData();
            }
        }
        
        function formatNumber(num) {
            return num.toString().replace(/\B(?=(\d{3})+(?!\d))/g, ",");
        }
        
        function somToUsd(som) {
            return (som / usdRate).toFixed(2);
        }
        
        async function loadData() {
            const period = document.getElementById('period').value;
            let url = '/api/expenses';
            let statsUrl = '/api/stats?period=' + period;
            
            if (period === 'custom') {
                const startDate = document.getElementById('startDate').value;
                const endDate = document.getElementById('endDate').value;
                if (startDate && endDate) {
                    url += `?start_date=${startDate}&end_date=${endDate}`;
                    statsUrl = '/api/stats?period=all';
                }
            }
            
            try {
                const [expensesRes, statsRes] = await Promise.all([
                    fetch(url),
                    fetch(statsUrl)
                ]);
                
                const expenses = await expensesRes.json();
                const stats = await statsRes.json();
                
                document.getElementById('totalAmount').textContent = formatNumber(stats.total) + ' som';
                document.getElementById('totalUSD').textContent = '$' + somToUsd(stats.total);
                document.getElementById('totalCount').textContent = stats.count + ' ta';
                
                const tbody = document.getElementById('tableBody');
                tbody.innerHTML = '';
                
                expenses.forEach((exp, idx) => {
                    const row = tbody.insertRow();
                    row.innerHTML = `
                        <td>${idx + 1}</td>
                        <td>${exp.date}</td>
                        <td>${exp.category}</td>
                        <td class="amount">${formatNumber(exp.amount)} som</td>
                        <td class="amount">$${somToUsd(exp.amount)}</td>
                        <td>${exp.description}</td>
                    `;
                });
                
                if (chartInstance) {
                    chartInstance.destroy();
                }
                
                const ctx = document.getElementById('categoryChart').getContext('2d');
                chartInstance = new Chart(ctx, {
                    type: 'bar',
                    data: {
                        labels: stats.categories.map(c => c.name),
                        datasets: [{
                            label: 'Summa (som)',
                            data: stats.categories.map(c => c.amount),
                            backgroundColor: 'rgba(102, 126, 234, 0.8)',
                            borderColor: 'rgba(102, 126, 234, 1)',
                            borderWidth: 2
                        }]
                    },
                    options: {
                        responsive: true,
                        plugins: {
                            title: {
                                display: true,
                                text: 'Kategoriyalar bo\'yicha xarajatlar',
                                font: { size: 18 }
                            }
                        },
                        scales: {
                            y: {
                                beginAtZero: true,
                                ticks: {
                                    callback: function(value) {
                                        return formatNumber(value) + ' som';
                                    }
                                }
                            }
                        }
                    }
                });
                
            } catch (error) {
                console.error('Xato:', error);
                alert('Malumotlarni yuklashda xato!');
            }
        }
        
        function exportExcel() {
            const period = document.getElementById('period').value;
            let url = '/api/export';
            
            if (period === 'custom') {
                const startDate = document.getElementById('startDate').value;
                const endDate = document.getElementById('endDate').value;
                if (startDate && endDate) {
                    url += `?start_date=${startDate}&end_date=${endDate}`;
                }
            }
            
            window.location.href = url;
        }
        
        loadCurrency();
        loadData();
    </script>
</body>
</html>
"""

if __name__ == "__main__":
    print("=" * 50)
    print("🌐 ADMIN PANEL SERVERI ISHGA TUSHDI!")
    print("=" * 50)
    print("\n📍 URL: http://localhost:5000")
    print(f"\n🔐 Parol: {ADMIN_PASSWORD}")
    print("✅ Session: 24 soat")
    print("\n⚠️  To'xtatish: Ctrl+C\n")
    print("=" * 50)
    app.run(host="0.0.0.0", port=5000, debug=True)
