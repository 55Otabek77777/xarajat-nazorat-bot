import asyncio
import logging
import os
from datetime import datetime, timedelta
from aiogram import Bot, Dispatcher, F
from aiogram.filters import Command
from aiogram.types import Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton, FSInputFile
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage

from config import BOT_TOKEN, ADMIN_PANEL_URL
from database import init_db, add_expense, get_all_expenses, get_expenses_by_date, get_expenses_by_date_range
from ai_helper import analyze_check_image, transcribe_voice

# Logging
logging.basicConfig(level=logging.INFO)

# Bot initialization
bot = Bot(token=BOT_TOKEN, parse_mode="HTML")
storage = MemoryStorage()
dp = Dispatcher(storage=storage)

# FSM States
class ExpenseForm(StatesGroup):
    waiting_for_amount = State()
    waiting_for_category = State()
    waiting_for_description = State()

# ==================== START ====================

@dp.message(Command("start"))
async def cmd_start(message: Message):
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📸 Chek rasmi", callback_data="photo")],
        [InlineKeyboardButton(text="🎤 Ovozli xabar", callback_data="voice")],
        [InlineKeyboardButton(text="✍️ Matn yozish", callback_data="text")],
        [InlineKeyboardButton(text="📊 Hisobotlar", callback_data="reports")],
        [InlineKeyboardButton(text="🔙 Orqaga", callback_data="back")]
    ])
    
    await message.answer(
        "✅ <b>BOT ISHGA TUSHDI!</b>\n\n"
        "=============================================\n\n"
        "📋 <b>Baza tayyorl</b>\n"
        "✅ <b>Ovozli xabar + AI tahlil</b>\n"
        "✅ <b>Chek rasmi + AI tahlil</b>\n"
        "✅ <b>Matn orqali kiritish</b>\n"
        "✅ <b>Hisobotlar (oylar, kunlar, kunlik)</b>\n"
        "✅ <b>Excel export (hafta/oy/yil)</b>\n"
        "✅ <b>Smart AI Fallback (10+ model)</b>\n\n"
        "=============================================\n\n"
        "Xarajat kiritish usulini tanlang:",
        reply_markup=keyboard
    )

@dp.callback_query(F.data == "start")
async def callback_start(callback: CallbackQuery):
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📸 Chek rasmi", callback_data="photo")],
        [InlineKeyboardButton(text="🎤 Ovozli xabar", callback_data="voice")],
        [InlineKeyboardButton(text="✍️ Matn yozish", callback_data="text")],
        [InlineKeyboardButton(text="📊 Hisobotlar", callback_data="reports")],
        [InlineKeyboardButton(text="🔙 Orqaga", callback_data="back")]
    ])
    
    await callback.message.edit_text(
        "Xarajat kiritish usulini tanlang:",
        reply_markup=keyboard
    )

# ==================== CHEK RASMI ====================

@dp.callback_query(F.data == "photo")
async def ask_for_photo(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("📸 Chek rasmini yuboring:")
    await state.set_state("waiting_for_photo")
    await callback.answer()

@dp.message(F.photo)
async def handle_photo(message: Message, state: FSMContext):
    current_state = await state.get_state()
    
    if current_state != "waiting_for_photo":
        return
    
    await message.answer("🔍 Chek tahlil qilinmoqda...")
    
    # Rasmni yuklab olish
    photo = message.photo[-1]
    file = await bot.get_file(photo.file_id)
    file_path = f"/tmp/{photo.file_id}.jpg"
    await bot.download_file(file.file_path, file_path)
    
    # AI tahlil
    result = await analyze_check_image(file_path)
    
    if "error" in result:
        await message.answer(f"❌ {result['error']}")
        await state.clear()
        return
    
    # Xarajatlarni saqlash
    if result.get("items"):
        for item in result["items"]:
            add_expense(
                amount=item.get("price", 0),
                category="CHEK",
                description=item.get("name", "Nomsiz mahsulot")
            )
        
        await message.answer(
            f"✅ Chek saqlandi!\n\n"
            f"💰 Jami: {result.get('total', 0):,.0f} so'm\n"
            f"📦 Mahsulotlar: {len(result['items'])} ta"
        )
    else:
        await message.answer("⚠️ Chekda ma'lumot topilmadi. Iltimos, qayta urinib ko'ring.")
    
    await state.clear()

# ==================== OVOZLI XABAR ====================

@dp.callback_query(F.data == "voice")
async def ask_for_voice(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("🎤 Ovozli xabar yuboring:")
    await state.set_state("waiting_for_voice")
    await callback.answer()

@dp.message(F.voice)
async def handle_voice(message: Message, state: FSMContext):
    current_state = await state.get_state()
    
    if current_state != "waiting_for_voice":
        return
    
    await message.answer("🔊 Ovoz tanilmoqda...")
    
    # Ovozni yuklab olish
    voice = message.voice
    file = await bot.get_file(voice.file_id)
    file_path = f"/tmp/{voice.file_id}.ogg"
    await bot.download_file(file.file_path, file_path)
    
    # AI tahlil
    result = await transcribe_voice(file_path)
    
    if "error" in result:
        await message.answer(f"❌ {result['error']}")
        await state.clear()
        return
    
    # Xarajatni saqlash
    if result.get("amount", 0) > 0:
        add_expense(
            amount=result["amount"],
            category=result.get("category", "BOSHQA"),
            description=result.get("transcription", "")
        )
        
        await message.answer(
            f"✅ Ovoz saqlandi!\n\n"
            f"💰 Summa: {result['amount']:,.0f} so'm\n"
            f"📁 Kategoriya: {result['category']}\n"
            f"📝 Matn: {result['transcription']}"
        )
    else:
        await message.answer("⚠️ Ovozda summa aniqlanmadi. Iltimos, matn orqali kiriting.")
    
    await state.clear()

# ==================== MATN YOZISH ====================

@dp.callback_query(F.data == "text")
async def ask_for_amount(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("💰 Xarajat summasini kiriting (faqat raqam):")
    await state.set_state(ExpenseForm.waiting_for_amount)
    await callback.answer()

@dp.message(ExpenseForm.waiting_for_amount)
async def process_amount(message: Message, state: FSMContext):
    try:
        amount = int(message.text.replace(" ", "").replace(",", ""))
        await state.update_data(amount=amount)
        
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🏗️ QURILISH", callback_data="cat_QURILISH")],
            [InlineKeyboardButton(text="🍎 OZIQ-OVQAT", callback_data="cat_OZIQ-OVQAT")],
            [InlineKeyboardButton(text="🚗 TRANSPORT", callback_data="cat_TRANSPORT")],
            [InlineKeyboardButton(text="📱 ALOQA", callback_data="cat_ALOQA")],
            [InlineKeyboardButton(text="📦 BOSHQA", callback_data="cat_BOSHQA")]
        ])
        
        await message.answer("📁 Kategoriyani tanlang:", reply_markup=keyboard)
        await state.set_state(ExpenseForm.waiting_for_category)
        
    except ValueError:
        await message.answer("❌ Noto'g'ri format! Faqat raqam kiriting (masalan: 50000)")

@dp.callback_query(F.data.startswith("cat_"))
async def process_category(callback: CallbackQuery, state: FSMContext):
    category = callback.data.replace("cat_", "")
    await state.update_data(category=category)
    
    await callback.message.edit_text("📝 Izoh yozing (yoki /skip buyrug'ini yuboring):")
    await state.set_state(ExpenseForm.waiting_for_description)
    await callback.answer()

@dp.message(ExpenseForm.waiting_for_description)
async def process_description(message: Message, state: FSMContext):
    description = message.text if message.text != "/skip" else ""
    
    data = await state.get_data()
    
    add_expense(
        amount=data["amount"],
        category=data["category"],
        description=description
    )
    
    await message.answer(
        f"✅ Xarajat saqlandi!\n\n"
        f"💰 Summa: {data['amount']:,.0f} so'm\n"
        f"📁 Kategoriya: {data['category']}\n"
        f"📝 Izoh: {description or '-'}"
    )
    
    await state.clear()

# ==================== HISOBOTLAR ====================

@dp.callback_query(F.data == "reports")
async def show_reports_menu(callback: CallbackQuery):
    """Hisobotlar asosiy menyusi"""
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📄 Haftalik hisobot", callback_data="report_week_text")],
        [InlineKeyboardButton(text="📄 Oylik hisobot", callback_data="report_month_text")],
        [InlineKeyboardButton(text="📄 Yillik hisobot", callback_data="report_year_text")],
        [InlineKeyboardButton(text="📅 Oylar bo'yicha", callback_data="report_by_months")],
        [InlineKeyboardButton(text="📊 Excel yuklash", callback_data="excel_menu")],
        [InlineKeyboardButton(text="🔙 Orqaga", callback_data="start")]
    ])
    
    await callback.message.edit_text(
        "📊 <b>Hisobotlar bo'limi</b>\n\n"
        "Kerakli hisobot turini tanlang:",
        reply_markup=keyboard
    )

# ========== MATN HISOBOTLAR ==========

@dp.callback_query(F.data == "report_week_text")
async def report_week_text(callback: CallbackQuery):
    """Haftalik hisobot (matn)"""
    today = datetime.now()
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    
    expenses = get_expenses_by_date_range(
        week_start.strftime("%Y-%m-%d"),
        week_end.strftime("%Y-%m-%d")
    )
    
    if not expenses:
        await callback.answer("⚠️ Bu haftada xarajat qilmadingiz!", show_alert=True)
        return
    
    category_totals = {}
    total_amount = 0
    
    report_text = f"📄 <b>Haftalik hisobot</b>\n"
    report_text += f"📅 {week_start.strftime('%d.%m.%Y')} - {week_end.strftime('%d.%m.%Y')}\n\n"
    
    for exp in expenses:
        category = exp[3]
        amount = exp[2]
        total_amount += amount
        
        if category not in category_totals:
            category_totals[category] = 0
        category_totals[category] += amount
    
    for category, amount in category_totals.items():
        report_text += f"• {category}: {amount:,.0f} so'm\n"
    
    report_text += f"\n💰 <b>Jami: {total_amount:,.0f} so'm</b>"
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔙 Orqaga", callback_data="reports")]
    ])
    
    await callback.message.edit_text(report_text, reply_markup=keyboard)

@dp.callback_query(F.data == "report_month_text")
async def report_month_text(callback: CallbackQuery):
    """Oylik hisobot (matn)"""
    today = datetime.now()
    month_start = today.replace(day=1)
    next_month = month_start + timedelta(days=32)
    month_end = next_month.replace(day=1) - timedelta(days=1)
    
    expenses = get_expenses_by_date_range(
        month_start.strftime("%Y-%m-%d"),
        month_end.strftime("%Y-%m-%d")
    )
    
    if not expenses:
        await callback.answer("⚠️ Bu oyda xarajat qilmadingiz!", show_alert=True)
        return
    
    category_totals = {}
    total_amount = 0
    
    month_names = ["", "Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun",
                   "Iyul", "Avgust", "Sentabr", "Oktabr", "Noyabr", "Dekabr"]
    
    report_text = f"📄 <b>Oylik hisobot</b>\n"
    report_text += f"📅 {month_names[today.month]} {today.year}\n\n"
    
    for exp in expenses:
        category = exp[3]
        amount = exp[2]
        total_amount += amount
        
        if category not in category_totals:
            category_totals[category] = 0
        category_totals[category] += amount
    
    for category, amount in category_totals.items():
        report_text += f"• {category}: {amount:,.0f} so'm\n"
    
    report_text += f"\n💰 <b>Jami: {total_amount:,.0f} so'm</b>"
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔙 Orqaga", callback_data="reports")]
    ])
    
    await callback.message.edit_text(report_text, reply_markup=keyboard)

@dp.callback_query(F.data == "report_year_text")
async def report_year_text(callback: CallbackQuery):
    """Yillik hisobot (matn)"""
    today = datetime.now()
    year_start = today.replace(month=1, day=1)
    year_end = today.replace(month=12, day=31)
    
    expenses = get_expenses_by_date_range(
        year_start.strftime("%Y-%m-%d"),
        year_end.strftime("%Y-%m-%d")
    )
    
    if not expenses:
        await callback.answer("⚠️ Bu yilda xarajat qilmadingiz!", show_alert=True)
        return
    
    category_totals = {}
    total_amount = 0
    
    report_text = f"📄 <b>Yillik hisobot</b>\n"
    report_text += f"📅 {today.year}-yil\n\n"
    
    for exp in expenses:
        category = exp[3]
        amount = exp[2]
        total_amount += amount
        
        if category not in category_totals:
            category_totals[category] = 0
        category_totals[category] += amount
    
    for category, amount in category_totals.items():
        report_text += f"• {category}: {amount:,.0f} so'm\n"
    
    report_text += f"\n💰 <b>Jami: {total_amount:,.0f} so'm</b>"
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔙 Orqaga", callback_data="reports")]
    ])
    
    await callback.message.edit_text(report_text, reply_markup=keyboard)

# ========== OYLAR BO'YICHA ==========

@dp.callback_query(F.data == "report_by_months")
async def show_months(callback: CallbackQuery):
    """12 oyni ko'rsatish"""
    months = [
        ("Yanvar", 1), ("Fevral", 2), ("Mart", 3),
        ("Aprel", 4), ("May", 5), ("Iyun", 6),
        ("Iyul", 7), ("Avgust", 8), ("Sentabr", 9),
        ("Oktabr", 10), ("Noyabr", 11), ("Dekabr", 12)
    ]
    
    keyboard = []
    for i in range(0, 12, 3):
        row = []
        for j in range(3):
            if i + j < 12:
                month_name, month_num = months[i + j]
                row.append(InlineKeyboardButton(
                    text=month_name,
                    callback_data=f"month_{month_num}"
                ))
        keyboard.append(row)
    
    keyboard.append([InlineKeyboardButton(text="🔙 Orqaga", callback_data="reports")])
    
    await callback.message.edit_text(
        "📅 <b>Oyni tanlang:</b>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=keyboard)
    )

@dp.callback_query(F.data.startswith("month_"))
async def show_days(callback: CallbackQuery):
    """Tanlangan oy kunlarini ko'rsatish"""
    month = int(callback.data.split("_")[1])
    year = datetime.now().year
    
    month_names = ["", "Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun",
                   "Iyul", "Avgust", "Sentabr", "Oktabr", "Noyabr", "Dekabr"]
    
    if month == 12:
        next_month = datetime(year + 1, 1, 1)
    else:
        next_month = datetime(year, month + 1, 1)
    
    last_day = (next_month - timedelta(days=1)).day
    
    keyboard = []
    row = []
    for day in range(1, last_day + 1):
        row.append(InlineKeyboardButton(
            text=str(day),
            callback_data=f"day_{year}_{month}_{day}"
        ))
        
        if len(row) == 7:
            keyboard.append(row)
            row = []
    
    if row:
        keyboard.append(row)
    
    keyboard.append([InlineKeyboardButton(text="🔙 Orqaga", callback_data="report_by_months")])
    
    await callback.message.edit_text(
        f"📅 <b>{month_names[month]} {year}</b>\n\n"
        "Kunni tanlang:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=keyboard)
    )

@dp.callback_query(F.data.startswith("day_"))
async def show_day_report(callback: CallbackQuery):
    """Tanlangan kun hisoboti"""
    parts = callback.data.split("_")
    year, month, day = int(parts[1]), int(parts[2]), int(parts[3])
    
    date_str = f"{year}-{month:02d}-{day:02d}"
    expenses = get_expenses_by_date(date_str)
    
    if not expenses:
        await callback.answer(f"⚠️ {day:02d}.{month:02d}.{year} kuni xarajat qilmadingiz!", show_alert=True)
        return
    
    report_text = f"📅 <b>{day:02d}.{month:02d}.{year} - Kunlik hisobot</b>\n\n"
    
    total = 0
    for exp in expenses:
        exp_id, date, amount, category, description = exp
        total += amount
        
        desc_text = f" ({description})" if description else ""
        report_text += f"• {category}: {amount:,.0f} so'm{desc_text}\n"
    
    report_text += f"\n💰 <b>Jami: {total:,.0f} so'm</b>"
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔙 Orqaga", callback_data=f"month_{month}")]
    ])
    
    await callback.message.edit_text(report_text, reply_markup=keyboard)

# ========== EXCEL YUKLASH ==========

@dp.callback_query(F.data == "excel_menu")
async def excel_menu(callback: CallbackQuery):
    """Excel yuklash menyusi"""
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📊 Haftalik Excel", callback_data="excel_week")],
        [InlineKeyboardButton(text="📊 Oylik Excel", callback_data="excel_month")],
        [InlineKeyboardButton(text="📊 Yillik Excel", callback_data="excel_year")],
        [InlineKeyboardButton(text="🔙 Orqaga", callback_data="reports")]
    ])
    
    await callback.message.edit_text(
        "📊 <b>Excel yuklash</b>\n\n"
        "Davr turini tanlang:",
        reply_markup=keyboard
    )

@dp.callback_query(F.data.startswith("excel_"))
async def generate_excel(callback: CallbackQuery):
    """Excel fayl yaratish"""
    period = callback.data.split("_")[1]
    
    today = datetime.now()
    
    if period == "week":
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)
        title = f"Haftalik hisobot ({start_date.strftime('%d.%m.%Y')}-{end_date.strftime('%d.%m.%Y')})"
    elif period == "month":
        start_date = today.replace(day=1)
        next_month = start_date + timedelta(days=32)
        end_date = next_month.replace(day=1) - timedelta(days=1)
        month_names = ["", "Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun",
                       "Iyul", "Avgust", "Sentabr", "Oktabr", "Noyabr", "Dekabr"]
        title = f"Oylik hisobot ({month_names[today.month]} {today.year})"
    else:
        start_date = today.replace(month=1, day=1)
        end_date = today.replace(month=12, day=31)
        title = f"Yillik hisobot ({today.year}-yil)"
    
    expenses = get_expenses_by_date_range(
        start_date.strftime("%Y-%m-%d"),
        end_date.strftime("%Y-%m-%d")
    )
    
    if not expenses:
        await callback.answer(f"⚠️ Bu davrda xarajat qilmadingiz!", show_alert=True)
        return
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Hisobot"
    
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    headers = ['Nomi', 'Sana', 'Turi', 'Izoh', "Summa (so'm)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    total = 0
    for row_num, exp in enumerate(expenses, start=4):
        exp_id, date, amount, category, description = exp
        
        ws.cell(row=row_num, column=1, value=category)
        ws.cell(row=row_num, column=2, value=date)
        ws.cell(row=row_num, column=3, value=category)
        ws.cell(row=row_num, column=4, value=description or "-")
        ws.cell(row=row_num, column=5, value=amount)
        
        total += amount
    
    last_row = len(expenses) + 4
    ws.cell(row=last_row, column=4, value="JAMI:").font = Font(bold=True)
    ws.cell(row=last_row, column=5, value=total).font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    
    filename = f"/tmp/hisobot_{period}_{today.strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    
    document = FSInputFile(filename)
    await callback.message.answer_document(
        document,
        caption=f"📊 {title}"
    )
    
    await callback.answer("✅ Excel fayl yuborildi!")

# ==================== ORQAGA ====================

@dp.callback_query(F.data == "back")
async def go_back(callback: CallbackQuery):
    await callback.answer("🔙 Asosiy menyuga qaytildi")
    await callback_start(callback)

# ==================== MAIN ====================

async def main():
    init_db()
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())